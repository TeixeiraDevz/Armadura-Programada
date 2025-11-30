#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar Inventario_ICs.xlsx a partir do CSV
Requisitos: pip install openpyxl
Execute: python gerar_xlsx.py
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import csv
    import os
except ImportError:
    print("ERRO: Biblioteca openpyxl não instalada.")
    print("Instale com: pip install openpyxl")
    exit(1)

# Dados do inventário
csv_file = 'Inventario_ICs.csv'
xlsx_file = 'Inventario_ICs.xlsx'

# Cores para formatação
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Inventário de ICs"

# Ler CSV e escrever no Excel
if os.path.exists(csv_file):
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border_style
                
                # Formatar cabeçalho
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 15,  # ID do IC
        'B': 50,  # Nome
        'C': 25,  # Tipo
        'D': 10,  # Versão
        'E': 25,  # Responsável
        'F': 80   # Localização
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'
    
    # Salvar arquivo
    wb.save(xlsx_file)
    print(f"✅ Arquivo {xlsx_file} gerado com sucesso!")
else:
    print(f"❌ Arquivo {csv_file} não encontrado!")
    print("Certifique-se de executar o script na pasta documentacao/ICs/")

